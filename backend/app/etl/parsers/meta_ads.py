"""Parser do export Meta Ads Manager (XLSX).

Estratégia:
- Lê headers por NOME (não por índice) — robusto a mudanças de ordem
- Detecta período via colunas "Início dos relatórios" / "Encerramento"
- Filtra automaticamente campanhas com impressões = 0 (irrelevantes)
- UPSERT por (ano, mes, nome_campanha) — pode subir XLSX múltiplas vezes
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mkt import MetaAdsCampanha


HEADER_MAP = {
    "Nome da campanha": "nome_campanha",
    "Veiculação da campanha": "veiculacao",
    "Orçamento do conjunto de anúncios": "orcamento_diario",
    "Valor usado (BRL)": "investimento",
    "Impressões": "impressoes",
    "Alcance": "alcance",
    "Cliques no link": "cliques",
    "CPM (custo por 1.000 impressões) (BRL)": "cpm",
    "CPC (custo por clique no link) (BRL)": "cpc",
    "CTR (taxa de cliques no link)": "ctr",
    "Frequência": "frequencia",
    "Resultados": "resultados",
    "Indicador de resultados": "indicador_resultado",
    "Custo por resultados": "custo_por_resultado",
    "Leads": "leads",
    "leads_imersão": "leads_imersao",
    "Compras/adição ao carrinho": "compras",
    "Valor dos resultados": "valor_resultados",
}


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _to_int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return 0


def _parse_xlsx(content: bytes) -> tuple[list[dict], int, int]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    periodo_inicio = None
    periodo_fim = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            periodo_inicio = row[0]
            periodo_fim = row[1]
            break

    if not periodo_inicio:
        raise ValueError("Não foi possível detectar período da planilha")

    if isinstance(periodo_inicio, str):
        periodo_inicio = datetime.fromisoformat(periodo_inicio.split()[0])
    elif hasattr(periodo_inicio, "year"):
        pass
    else:
        raise ValueError(f"Formato de data inválido: {periodo_inicio!r}")

    ano = periodo_inicio.year
    if isinstance(periodo_fim, str):
        try:
            periodo_fim = datetime.fromisoformat(periodo_fim.split()[0])
        except Exception:
            periodo_fim = periodo_inicio
    mes = periodo_fim.month if hasattr(periodo_fim, "month") else periodo_inicio.month

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        impr_idx = header_idx.get("Impressões")
        if impr_idx is None:
            continue
        impr = _to_int(row[impr_idx])
        if impr == 0:
            continue

        d: dict[str, Any] = {"ano": ano, "mes": mes, "impressoes": impr}

        for header, col_name in HEADER_MAP.items():
            if col_name in d:
                continue
            idx = header_idx.get(header)
            if idx is None:
                continue
            val = row[idx]
            if col_name in (
                "investimento", "orcamento_diario", "cpm", "cpc", "ctr",
                "frequencia", "custo_por_resultado", "valor_resultados",
            ):
                d[col_name] = _to_decimal(val) or Decimal("0")
            elif col_name in (
                "alcance", "cliques", "resultados", "leads", "leads_imersao", "compras",
            ):
                d[col_name] = _to_int(val)
            else:
                d[col_name] = str(val) if val is not None else None

        if not d.get("nome_campanha"):
            continue

        d["nome_campanha"] = d["nome_campanha"][:500]
        rows.append(d)

    return rows, ano, mes


async def importar_meta_ads(content: bytes, db: AsyncSession) -> dict:
    rows, ano, mes = _parse_xlsx(content)

    if not rows:
        return {
            "rows_processed": 0,
            "rows_inserted": 0,
            "rows_updated": 0,
            "rows_skipped": 0,
            "warnings": ["Nenhuma campanha com impressões > 0 encontrada"],
            "period_detected": f"{ano}-{mes:02d}",
        }

    inserted = 0
    updated = 0

    for row in rows:
        stmt = pg_insert(MetaAdsCampanha).values(**row)
        update_cols = {
            k: stmt.excluded[k]
            for k in row.keys()
            if k not in ("ano", "mes", "nome_campanha")
        }
        update_cols["atualizado_em"] = text("now()")

        stmt = stmt.on_conflict_do_update(
            constraint="uq_meta_ads_periodo_campanha",
            set_=update_cols,
        )

        result = await db.execute(stmt.returning(MetaAdsCampanha.id, text("(xmax = 0) AS inserted")))
        row_result = result.first()
        if row_result and row_result[1]:
            inserted += 1
        else:
            updated += 1

    await db.commit()

    return {
        "rows_processed": len(rows),
        "rows_inserted": inserted,
        "rows_updated": updated,
        "rows_skipped": 0,
        "warnings": [],
        "period_detected": f"{ano}-{mes:02d}",
    }
