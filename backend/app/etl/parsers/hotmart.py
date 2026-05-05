"""Parser do export Hotmart Sales History.

Estratégia:
- Aceita .xls (que é xlsx mascarado) e .xlsx
- UPSERT por transacao
- Mantém TODAS as vendas (Aprovado + Completo + outros) — frontend filtra depois
"""
from __future__ import annotations

import io
import warnings
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mkt import VendaHotmart


HEADER_MAP = {
    "Transação": "transacao",
    "Nome do Produto": "produto",
    "Nome do Produtor": "produtor",
    "Nome do Afiliado": "afiliado",
    "Meio de Pagamento": "meio_pagamento",
    "Moeda": "moeda",
    "Preço Total": "preco_total",
    "Faturamento líquido": "faturamento_liquido",
    "Número da Parcela": "numero_parcela",
    "Recorrência": "recorrencia",
    "Data de Venda": "data_venda",
    "Data de Confirmação": "data_confirmacao",
    "Status": "status",
    "Nome": "cliente_nome",
    "Email": "cliente_email",
    "Estado": "cliente_estado",
    "País": "cliente_pais",
    "Código do Produto": "codigo_produto",
    "Código de Oferta": "codigo_oferta",
    "Tipo pagamento oferta": "tipo_pagamento_oferta",
}


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def _parse_dt(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if hasattr(v, "year"):
        return v if isinstance(v, datetime) else datetime.combine(v, datetime.min.time())
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.split(".")[0], fmt)
            except ValueError:
                continue
    return None


def _parse_xlsx(content: bytes) -> list[dict]:
    """Hotmart .xls é xlsx mascarado. Aceita ambos."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(io.BytesIO(content), data_only=True)

    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        d: dict[str, Any] = {}
        for header, col_name in HEADER_MAP.items():
            idx = header_idx.get(header)
            if idx is None:
                continue
            val = row[idx]
            if col_name in ("preco_total", "faturamento_liquido"):
                d[col_name] = _to_decimal(val)
            elif col_name == "numero_parcela":
                d[col_name] = _to_int(val)
            elif col_name in ("data_venda", "data_confirmacao"):
                d[col_name] = _parse_dt(val)
            else:
                d[col_name] = str(val)[:500] if val is not None else None

        if not d.get("transacao") or not d.get("produto"):
            continue
        d["transacao"] = d["transacao"][:100]
        d["produto"] = d["produto"][:500]

        rows.append(d)

    return rows


async def importar_hotmart(content: bytes, db: AsyncSession) -> dict:
    rows = _parse_xlsx(content)

    if not rows:
        return {
            "rows_processed": 0,
            "rows_inserted": 0,
            "rows_updated": 0,
            "rows_skipped": 0,
            "warnings": ["Nenhuma venda encontrada na planilha"],
            "period_detected": None,
        }

    inserted = 0
    updated = 0

    for row in rows:
        stmt = pg_insert(VendaHotmart).values(**row)
        update_cols = {
            k: stmt.excluded[k]
            for k in row.keys()
            if k != "transacao"
        }
        update_cols["atualizado_em"] = text("now()")
        stmt = stmt.on_conflict_do_update(
            constraint="uq_vendas_hotmart_transacao",
            set_=update_cols,
        )
        result = await db.execute(stmt.returning(VendaHotmart.id, text("(xmax = 0) AS inserted")))
        row_result = result.first()
        if row_result and row_result[1]:
            inserted += 1
        else:
            updated += 1

    await db.commit()

    periods: Counter = Counter()
    for r in rows:
        dt = r.get("data_venda")
        if dt:
            periods[(dt.year, dt.month)] += 1
    period_detected = None
    if periods:
        ano, mes = periods.most_common(1)[0][0]
        period_detected = f"{ano}-{mes:02d}"

    return {
        "rows_processed": len(rows),
        "rows_inserted": inserted,
        "rows_updated": updated,
        "rows_skipped": 0,
        "warnings": [],
        "period_detected": period_detected,
    }
