"""ETL pra DASH_Processo_Seletivo_2026.xlsx → comercial.funil_resultado.

Estrutura das abas Janeiro/Fevereiro/Março/Abril/Maio:
- Linha header: 'Mês' na col B + nomes de produtos nas cols C+
- Blocos: 'Leads Gerados', 'Ligação Qualificação', 'SQL/Reunião Agendada',
  'Reunião Realizadas', 'Alunos - Vendas' — cada um com Meta/Resultado.

Idempotente: UPSERT em (produto_id, etapa_id, ano, mes).

Resolução de produto:
  1. Match exato por nome
  2. Match normalizado (lower + sem acento)
  3. Cria com tipo='pos_graduacao' (todas as cols das abas mensais são pós-grad)

Uso:
    uv run python -m app.etl.parsers.comercial \\
      --planilha /opt/dashcenat/data/DASH_Processo_Seletivo_2026.xlsx \\
      --ano 2026
"""
import argparse
import asyncio
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core.db import async_session
from app.etl.parsers._helpers import normalize, to_decimal
from app.models.catalogo import Produto
from app.models.comercial import FunilEtapa, FunilResultado


MES_NOME_PARA_NUM = {
    "Janeiro": 1, "Fevereiro": 2, "Marco": 3, "Março": 3,
    "Abril": 4, "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8,
    "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12,
}

BLOCO_PARA_ETAPA = {
    "leads gerados": 1,
    "ligacao qualificacao": 2,
    "sql / reuniao agendada": 3,
    "sql/reuniao agendada": 3,
    "reuniao realizadas": 4,
    "reuniao realizada": 4,
    "reunioes realizadas": 4,
    "alunos - vendas": 5,
    "alunos-vendas": 5,
    "alunos vendas": 5,
}

ABAS_ALVO = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio"]


def is_label(label: Any, expected: str) -> bool:
    return isinstance(label, str) and normalize(label) == expected


async def resolver_produto(db, nome_coluna: str, cache: dict):
    nome_norm = normalize(nome_coluna)
    if nome_norm in cache:
        return cache[nome_norm]

    # Match exato
    p = (await db.execute(select(Produto).where(Produto.nome == nome_coluna))).scalar_one_or_none()
    if p:
        cache[nome_norm] = p
        return p

    # Match normalizado
    for prod in (await db.execute(select(Produto))).scalars().all():
        if normalize(prod.nome) == nome_norm:
            cache[nome_norm] = prod
            return prod

    # Cria
    print(f"  ➕ criando produto: '{nome_coluna}' (tipo=pos_graduacao)")
    novo = Produto(tipo="pos_graduacao", nome=nome_coluna.strip(), codigo=None)
    db.add(novo)
    await db.flush()
    cache[nome_norm] = novo
    return novo


def parse_aba(ws) -> list[tuple]:
    """Retorna lista de (etapa_id, col_idx, nome_produto, meta, resultado)."""
    rows = list(ws.iter_rows(values_only=False))

    # Encontra header (linha onde col B é 'Mês' ou 'Data de atualização...')
    header_idx = None
    for i, row in enumerate(rows):
        b = row[1].value if len(row) > 1 else None
        if isinstance(b, str) and normalize(b) in ("mes", "data de atualizacao da planilha"):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 3  # fallback

    header = rows[header_idx]

    # Produtos das cols C+ até encontrar TOTAL ou vazio
    produtos_cols = []
    for col_idx in range(2, len(header)):
        val = header[col_idx].value
        if val and isinstance(val, str):
            v = val.strip()
            n = normalize(v)
            if n in ("total", "") or v.startswith("Data de") or len(v) < 2:
                break
            produtos_cols.append((col_idx, v))
        elif val is None and produtos_cols:
            break

    print(f"    Produtos detectados: {len(produtos_cols)}")
    for col, nome in produtos_cols:
        print(f"      col={col}, nome='{nome}'")

    # Itera blocos
    out = []
    i = header_idx + 1
    while i < len(rows):
        row = rows[i]
        label = row[1].value if len(row) > 1 else None
        if isinstance(label, str):
            etapa_id = BLOCO_PARA_ETAPA.get(normalize(label))
            if etapa_id:
                meta_row = rows[i + 1] if i + 1 < len(rows) else None
                resultado_row = rows[i + 2] if i + 2 < len(rows) else None
                if (meta_row and is_label(meta_row[1].value, "meta")
                    and resultado_row and is_label(resultado_row[1].value, "resultado")):
                    for col_idx, nome_prod in produtos_cols:
                        meta = to_decimal(meta_row[col_idx].value if col_idx < len(meta_row) else None)
                        resultado = to_decimal(resultado_row[col_idx].value if col_idx < len(resultado_row) else None)
                        out.append((etapa_id, col_idx, nome_prod, meta, resultado))
                    i += 3
                    continue
        i += 1

    return out


async def importar(planilha_path: str, ano: int):
    print(f"=== ETL Comercial: {planilha_path} (ano={ano}) ===")
    wb = load_workbook(planilha_path, read_only=False, data_only=True)
    sheets = [s for s in wb.sheetnames if s in ABAS_ALVO]
    print(f"Abas a processar: {sheets}")

    total_inseridos = 0
    total_skipped = 0
    cache_produtos: dict = {}

    async with async_session() as db:
        # Garante etapas seedadas
        etapas = (await db.execute(select(FunilEtapa))).scalars().all()
        if len(etapas) < 5:
            raise RuntimeError("Etapas do funil não estão seedadas. Rode 'python -m app.etl.seed' primeiro.")

        for sheet_name in sheets:
            mes = MES_NOME_PARA_NUM.get(sheet_name)
            if not mes:
                print(f"  ⚠️ Aba '{sheet_name}' sem mês — skip")
                continue

            print(f"\n  📅 Processando '{sheet_name}' (mês={mes})...")
            tuplas = parse_aba(wb[sheet_name])
            print(f"    {len(tuplas)} entradas (etapa × produto)")

            for etapa_id, col_idx, nome_prod, meta, resultado in tuplas:
                if meta is None and resultado is None:
                    total_skipped += 1
                    continue
                produto = await resolver_produto(db, nome_prod, cache_produtos)
                stmt = pg_insert(FunilResultado).values(
                    produto_id=produto.id, etapa_id=etapa_id,
                    ano=ano, mes=mes, meta=meta, resultado=resultado,
                ).on_conflict_do_update(
                    index_elements=["produto_id", "etapa_id", "ano", "mes"],
                    set_={"meta": meta, "resultado": resultado, "atualizado_em": text("NOW()")},
                )
                await db.execute(stmt)
                total_inseridos += 1

        await db.commit()

    wb.close()

    print("\n=== Concluído ===")
    print(f"Inseridos/atualizados: {total_inseridos}")
    print(f"Skipped (meta+resultado nulos): {total_skipped}")
    print(f"Produtos no cache: {len(cache_produtos)}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--planilha", required=True)
    parser.add_argument("--ano", type=int, required=True)
    args = parser.parse_args()
    if not Path(args.planilha).is_file():
        print(f"❌ Arquivo não encontrado: {args.planilha}", file=sys.stderr)
        sys.exit(1)
    asyncio.run(importar(args.planilha, args.ano))


if __name__ == "__main__":
    main()
