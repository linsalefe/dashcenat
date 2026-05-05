"""Parser de lançamento (a partir de XLSX simples ou da aba 'Lançamento'
do Sistema_Marketing___Inscrições.xlsx).

A planilha do CENAT segue padrão linha "Meta" e linha "Resultado" pra
cada métrica. Este parser foca em estrutura simples e flexível —
aceita upload de .xlsx com colunas:
    nome | ano | mes | investimento_meta | investimento_resultado |
    leads_meta | leads_organico | leads_pago | cpl_meta | cpl_resultado |
    mqls_meta | mqls_resultado | alunos_meta | alunos_resultado |
    receita_meta | receita_resultado | engajamento_json
"""
from __future__ import annotations

import io
import json
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mkt import Lancamento


COLUNAS_OBRIGATORIAS = {"nome", "ano", "mes"}


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def _parse_engajamento(v: Any) -> dict:
    if not v:
        return {}
    if isinstance(v, dict):
        return v
    if isinstance(v, str):
        try:
            return json.loads(v)
        except (json.JSONDecodeError, ValueError):
            return {"raw": v}
    return {}


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    missing = COLUNAS_OBRIGATORIAS - set(header_idx.keys())
    if missing:
        raise ValueError(
            f"Colunas obrigatórias ausentes: {missing}. "
            f"Headers encontrados: {headers}"
        )

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[header_idx["nome"]]:
            continue

        d: dict[str, Any] = {
            "nome": str(row[header_idx["nome"]])[:200],
            "ano": _to_int(row[header_idx["ano"]]),
            "mes": _to_int(row[header_idx["mes"]]),
        }

        if d["ano"] is None or d["mes"] is None:
            continue

        for campo, fn in [
            ("investimento_meta", _to_decimal),
            ("investimento_resultado", _to_decimal),
            ("leads_meta", _to_int),
            ("leads_organico", _to_int),
            ("leads_pago", _to_int),
            ("leads_total", _to_int),
            ("cpl_meta", _to_decimal),
            ("cpl_resultado", _to_decimal),
            ("mqls_meta", _to_int),
            ("mqls_resultado", _to_int),
            ("alunos_meta", _to_int),
            ("alunos_resultado", _to_int),
            ("receita_meta", _to_decimal),
            ("receita_resultado", _to_decimal),
        ]:
            if campo in header_idx:
                v = fn(row[header_idx[campo]])
                if v is not None:
                    d[campo] = v

        if "engajamento_json" in header_idx:
            d["engajamento"] = _parse_engajamento(row[header_idx["engajamento_json"]])

        d.setdefault("leads_organico", 0)
        d.setdefault("leads_pago", 0)
        d.setdefault(
            "leads_total",
            (d.get("leads_organico") or 0) + (d.get("leads_pago") or 0),
        )

        rows.append(d)

    return rows


async def importar_lancamentos(content: bytes, db: AsyncSession) -> dict:
    rows = _parse_xlsx(content)

    if not rows:
        return {
            "rows_processed": 0,
            "rows_inserted": 0,
            "rows_updated": 0,
            "rows_skipped": 0,
            "warnings": ["Nenhum lançamento válido na planilha"],
            "period_detected": None,
        }

    inserted = 0
    updated = 0

    for row in rows:
        stmt = pg_insert(Lancamento).values(**row)
        update_cols = {
            k: stmt.excluded[k]
            for k in row.keys()
            if k not in ("ano", "mes", "nome")
        }
        update_cols["atualizado_em"] = text("now()")
        stmt = stmt.on_conflict_do_update(
            constraint="uq_lancamento_periodo_nome",
            set_=update_cols,
        )
        result = await db.execute(stmt.returning(Lancamento.id, text("(xmax = 0) AS inserted")))
        row_result = result.first()
        if row_result and row_result[1]:
            inserted += 1
        else:
            updated += 1

    await db.commit()

    if rows:
        ano, mes = rows[0]["ano"], rows[0]["mes"]
        period_detected = f"{ano}-{mes:02d}"
    else:
        period_detected = None

    return {
        "rows_processed": len(rows),
        "rows_inserted": inserted,
        "rows_updated": updated,
        "rows_skipped": 0,
        "warnings": [],
        "period_detected": period_detected,
    }
